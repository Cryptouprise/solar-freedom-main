"""
Keyword Ranking Report Builder — breakyoursolarcontract.com
Mirrors the exact format of the leadflowpartners.com report:
  - Single sheet "Keyword Status Report"
  - Column A: Keywords
  - Column B: Starting baseline (today)
  - Columns C+: Monthly snapshots (auto-added each month)
  - Color coding: Green (1-10), Yellow (11-20), Orange (21-50), Red (51-100), Gray (Not Found)
  - Footer: Search Engine + Target Location
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ─── Current GSC Data (pulled Mar 30, 2026 — 3-month window) ──────────────────
# Format: (keyword, avg_position_rounded)
# Source: Google Search Console → Performance → Queries (breakyoursolarcontract.com)
# Positions are averages over 3 months; "Not Found" = not in top 100

KEYWORDS = [
    # Page 1 of GSC (sorted by impressions desc)
    ("selling a house with leased solar panels", 65),
    ("how to close a solar deal", 41),
    ("cancel solar contract", 58),
    ("stuck in solar panel contract", 12),
    ("how to get out of a solar panel lease", 15),
    ("how to cancel solar panel contract", 51),
    ("solar panel contract", 57),
    ("can i cancel my solar contract", 9),
    ("how to get out of a solar contract", 22),
    ("how to cancel sunrun contract before installation", 26),
    # Page 2 of GSC
    ("how to cancel solar contract", 29),
    ("solar release co", 39),
    ("solar panel cancellation", 46),
    ("leasing solar panels selling house", 47),
    ("how to cancel sunrun contract", 48),
    ("how to get out of a solar lease", 49),
    ("selling house with leased solar panels", 50),
    ("get out of solar lease", "Not Found"),
    ("selling a home with ppa solar panels", "Not Found"),
    ("solar contracts", "Not Found"),
    # Additional target keywords (not yet ranking — priority targets)
    ("break solar contract", "Not Found"),
    ("solar contract cancellation", "Not Found"),
    ("cancel sunrun contract", "Not Found"),
    ("cancel vivint solar contract", "Not Found"),
    ("cancel tesla solar contract", "Not Found"),
    ("solar panel contract buyout", "Not Found"),
    ("how to get out of a sunrun contract", "Not Found"),
    ("solar lease cancellation", "Not Found"),
    ("solar loan cancellation", "Not Found"),
    ("solar panel contract problems", "Not Found"),
    ("solar company fraud", "Not Found"),
    ("solar panel scam", "Not Found"),
    ("sunrun complaints", "Not Found"),
    ("sunnova complaints", "Not Found"),
    ("vivint solar complaints", "Not Found"),
    ("solar contract attorney", "Not Found"),
    ("solar panel contract lawyer", "Not Found"),
    ("selling house with solar panels financed", "Not Found"),
    ("solar panel lien on house", "Not Found"),
    ("solar panel debt on home sale", "Not Found"),
]

# ─── Color Definitions ─────────────────────────────────────────────────────────
def get_fill(position):
    """Return PatternFill based on ranking position."""
    if position == "Not Found":
        return PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
    elif isinstance(position, (int, float)):
        pos = int(position)
        if pos <= 10:
            return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        elif pos <= 20:
            return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        elif pos <= 50:
            return PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Orange
        else:
            return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red/Pink
    return None

def get_font_color(position):
    """Return Font color based on ranking position."""
    if position == "Not Found":
        return Font(color="999999", size=10)
    elif isinstance(position, (int, float)):
        pos = int(position)
        if pos <= 10:
            return Font(color="276221", size=10, bold=True)   # Dark green
        elif pos <= 20:
            return Font(color="9C6500", size=10, bold=True)   # Dark yellow
        elif pos <= 50:
            return Font(color="974706", size=10, bold=True)   # Dark orange
        else:
            return Font(color="9C0006", size=10, bold=True)   # Dark red
    return Font(size=10)

# ─── Build Workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Keyword Status Report"

# Header fills
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
header_font = Font(color="FFFFFF", bold=True, size=10)
subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # Medium blue
subheader_font = Font(color="FFFFFF", bold=True, size=10)

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# ─── Row 1: Title Banner ───────────────────────────────────────────────────────
ws.merge_cells('A1:H1')
title_cell = ws['A1']
title_cell.value = "Keyword Ranking Report — breakyoursolarcontract.com"
title_cell.fill = PatternFill(start_color="0D0F14", end_color="0D0F14", fill_type="solid")
title_cell.font = Font(color="F97316", bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# ─── Row 2: Column Headers ─────────────────────────────────────────────────────
today = datetime(2026, 3, 30)
headers = [
    "Keywords",
    f"30-Mar-26 (Starting)",
    # Future monthly columns will be added here
]
# Future months placeholder headers (empty for now, to be filled monthly)
future_months = [
    datetime(2026, 4, 30),
    datetime(2026, 5, 31),
    datetime(2026, 6, 30),
    datetime(2026, 7, 31),
    datetime(2026, 8, 31),
    datetime(2026, 9, 30),
]

all_headers = ["Keywords", "30-Mar-26 (Starting)"] + [d.strftime("%-d-%b-%y") for d in future_months]

for col_idx, header in enumerate(all_headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.fill = header_fill if col_idx == 1 else subheader_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[2].height = 32

# ─── Column A width ────────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 48
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws.column_dimensions[col_letter].width = 16

# ─── Data Rows ─────────────────────────────────────────────────────────────────
for row_idx, (keyword, position) in enumerate(KEYWORDS, 3):
    # Column A — keyword name
    kw_cell = ws.cell(row=row_idx, column=1, value=keyword)
    kw_cell.font = Font(size=10)
    kw_cell.alignment = Alignment(vertical='center', wrap_text=False)
    kw_cell.border = thin_border
    # Alternate row shading for readability
    if row_idx % 2 == 0:
        kw_cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

    # Column B — starting position (today)
    pos_cell = ws.cell(row=row_idx, column=2, value=position)
    fill = get_fill(position)
    font = get_font_color(position)
    if fill:
        pos_cell.fill = fill
    pos_cell.font = font
    pos_cell.alignment = Alignment(horizontal='center', vertical='center')
    pos_cell.border = thin_border

    # Columns C-H — empty placeholders for future monthly snapshots
    for col_idx in range(3, 9):
        empty_cell = ws.cell(row=row_idx, column=col_idx, value="")
        empty_cell.border = thin_border
        empty_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[row_idx].height = 18

# ─── Legend Row ────────────────────────────────────────────────────────────────
legend_row = len(KEYWORDS) + 4
ws.merge_cells(f'A{legend_row}:B{legend_row}')
ws[f'A{legend_row}'].value = "COLOR LEGEND"
ws[f'A{legend_row}'].font = Font(bold=True, size=9, color="444444")
ws[f'A{legend_row}'].alignment = Alignment(horizontal='left')

legend_items = [
    (legend_row + 1, "C6EFCE", "276221", "Position 1–10 (Page 1 — Top Rankings)"),
    (legend_row + 2, "FFEB9C", "9C6500", "Position 11–20 (Page 2)"),
    (legend_row + 3, "FFCC99", "974706", "Position 21–50 (Pages 3–5)"),
    (legend_row + 4, "FFC7CE", "9C0006", "Position 51–100 (Deep pages)"),
    (legend_row + 5, "F2F2F2", "999999", "Not Found (Not in top 100)"),
]
for r, bg, fg, label in legend_items:
    ws.merge_cells(f'A{r}:B{r}')
    cell = ws[f'A{r}']
    cell.value = label
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[r].height = 16

# ─── Footer ────────────────────────────────────────────────────────────────────
footer_row = legend_row + 8
ws[f'A{footer_row}'].value = "Search Engine : Google.com"
ws[f'A{footer_row}'].font = Font(size=9, color="666666", italic=True)
ws[f'A{footer_row + 1}'].value = "Target Location : USA"
ws[f'A{footer_row + 1}'].font = Font(size=9, color="666666", italic=True)
ws[f'A{footer_row + 2}'].value = f"Report Generated : {today.strftime('%B %d, %Y')}"
ws[f'A{footer_row + 2}'].font = Font(size=9, color="666666", italic=True)
ws[f'A{footer_row + 3}'].value = "Data Source : Google Search Console (3-month window)"
ws[f'A{footer_row + 3}'].font = Font(size=9, color="666666", italic=True)

# ─── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = 'B3'

# ─── Save ─────────────────────────────────────────────────────────────────────
output_path = "/home/ubuntu/solar-freedom/docs/KeywordRankingReport_breakyoursolarcontract_30-Mar-2026.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Report saved to: {output_path}")
print(f"Total keywords tracked: {len(KEYWORDS)}")
ranking_count = sum(1 for _, p in KEYWORDS if p != "Not Found")
print(f"Currently ranking (in top 100): {ranking_count}")
page1_count = sum(1 for _, p in KEYWORDS if isinstance(p, int) and p <= 10)
print(f"On page 1 (positions 1-10): {page1_count}")
page2_count = sum(1 for _, p in KEYWORDS if isinstance(p, int) and 11 <= p <= 20)
print(f"On page 2 (positions 11-20): {page2_count}")
